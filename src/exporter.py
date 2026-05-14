"""
Export receipts to the official VAT Refund XLSX template
and to a PDF containing all receipt images.
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Frame,
    KeepInFrame,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from . import config, db

# XLSX layout (inspected from the template):
# Columns: A=#, B=Date, C=Vendor, D=Receipt#, E=VAT Amount
# D5:E5 merged  = employee name field (Table sheet) / =Table!D5 formula (continuation sheets)
# D6:E6 merged  = grand total (Table sheet only)
# E40            = per-sheet SUM (each sheet)
SHEETS_ORDER = ["Table", "Continuation", "Continuation2", "Continuation3"]
ROWS_PER_SHEET = 30
# How many receipts fit in one workbook (4 sheets × 30 rows). Years with
# more than this need to be split into multiple "copy N" workbooks.
RECEIPTS_PER_XLSX = len(SHEETS_ORDER) * ROWS_PER_SHEET
DATA_START_ROW = 9   # First data row (A9 = 1)
NAME_ROW = 5         # Row of the employee name merged cell D5:E5
NAME_COL = 4         # Column D
GRAND_TOTAL_ROW = 6  # Row of grand total merged cell D6:E6
GRAND_TOTAL_COL = 4  # Column D
PER_SHEET_TOTAL_ROW = 40  # Row with per-sheet SUM
PER_SHEET_TOTAL_COL = 5   # Column E


def _display_vendor(doc: dict) -> str:
    return doc.get("display_vendor") or doc.get("printed_vendor") or doc.get("vendor") or ""


def _find_data_start(ws) -> int:
    """Find the row where numbered data starts (cell A = 1)."""
    for row in range(1, 50):
        val = ws.cell(row=row, column=1).value
        if val == 1 or val == "1":
            return row
    return DATA_START_ROW


def _find_name_cell(ws) -> tuple[int, int] | None:
    """Find the 'Employee Name' label and return the cell to the right of it."""
    for row in range(1, 10):
        for col in range(1, 10):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and "Employee Name" in val:
                return (row, col + 1)
    return None


async def build_xlsx(telegram_id: int, employee_name: str, output_path: Path) -> Path:
    """Fill the VAT_Refund template with all of the user's receipts."""
    receipts = await db.list_receipts(telegram_id)
    return build_xlsx_from_receipts(receipts, employee_name, output_path)


def build_xlsx_from_receipts(
    receipts: list[dict],
    employee_name: str,
    output_path: Path,
    start_seq: int = 1,
) -> Path:
    """Render a pre-filtered list of receipts into a fresh copy of the template.
    Used by /export_vat to produce one workbook per calendar year.

    `start_seq` is the visible row number printed in column A of the first
    receipt. Defaults to 1 (a standalone workbook). When a year is split
    across multiple copies because it exceeds the 120-row template, copy 2
    passes start_seq=121 so the numbering stays continuous if the tax
    office staples the copies together."""
    shutil.copy(config.TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)

    cumulative_total = 0.0

    # Walk every sheet in the template — even empty ones get the employee
    # name so continuations stay identified. Per-sheet totals accumulate
    # (E40 on each sheet = running total through that sheet) so the user
    # can see the refund amount grow as they flip across tabs.
    for idx, sheet_name in enumerate(SHEETS_ORDER):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        chunk = receipts[idx * ROWS_PER_SHEET : (idx + 1) * ROWS_PER_SHEET]

        # Employee name on every continuation tab (D5 is the TL of D5:E5).
        ws.cell(row=NAME_ROW, column=NAME_COL, value=employee_name)

        # Write receipt rows. Column A (sequential #) is written explicitly
        # because the template's pre-filled numbers display with a broken
        # number format on some Excel versions.
        sheet_total = 0.0
        for i, r in enumerate(chunk):
            row = DATA_START_ROW + i
            seq = start_seq + idx * ROWS_PER_SHEET + i
            num_cell = ws.cell(row=row, column=1, value=seq)
            num_cell.number_format = "General"
            # Manual entries: bold the row number so the finance office can
            # spot rows that were typed in (and could contain typos) rather
            # than fetched and verified from soliq.uz.
            if r.get("manual"):
                existing = num_cell.font
                num_cell.font = Font(
                    name=existing.name, size=existing.size,
                    color=existing.color, family=existing.family,
                    bold=True,
                )
            ws.cell(row=row, column=2, value=r.get("date", ""))
            ws.cell(row=row, column=3, value=_display_vendor(r))
            ws.cell(row=row, column=4, value=r.get("receipt_number", ""))
            vat = float(r.get("vat_amount") or 0)
            ws.cell(row=row, column=5, value=vat)
            sheet_total += vat

        # Blank leftover rows in this sheet's data block so stale template
        # numbers (1-30) don't show through when the chunk is short/empty.
        for i in range(len(chunk), ROWS_PER_SHEET):
            row = DATA_START_ROW + i
            for col in range(1, 6):
                ws.cell(row=row, column=col).value = None

        # Running total through this sheet (E40), replacing the SUM formula.
        cumulative_total += sheet_total
        ws.cell(row=PER_SHEET_TOTAL_ROW, column=PER_SHEET_TOTAL_COL, value=cumulative_total)

    # Grand total on Table!D6 (D6:E6 merged cell).
    if "Table" in wb.sheetnames:
        wb["Table"].cell(row=GRAND_TOTAL_ROW, column=GRAND_TOTAL_COL, value=cumulative_total)

    wb.save(output_path)
    return output_path


# Each receipt is one rendered page (no embedded photos), so PDFs stay
# small. The cap is kept generous to limit cover-page bloat and keep a
# single year manageable for the tax office to review.
RECEIPTS_PER_PDF = 60


async def build_pdf(telegram_id: int, employee_name: str, output_path: Path) -> Path:
    """Back-compat single-file PDF builder. Prefer build_pdfs for large sets."""
    paths = await build_pdfs(telegram_id, employee_name, output_path.parent, output_path.stem)
    # When only one file was produced, rename it to match the requested name.
    if len(paths) == 1 and paths[0] != output_path:
        paths[0].rename(output_path)
        return output_path
    return paths[0]


async def build_pdfs(
    telegram_id: int,
    employee_name: str,
    output_dir: Path,
    name_prefix: str = "Receipts",
) -> list[Path]:
    """Generate one or more PDFs (split every RECEIPTS_PER_PDF receipts)."""
    receipts = await db.list_receipts(telegram_id)
    return await build_pdfs_from_receipts(
        receipts, employee_name, output_dir, name_prefix
    )


async def build_pdfs_from_receipts(
    receipts: list[dict],
    employee_name: str,
    output_dir: Path,
    name_prefix: str = "Receipts",
) -> list[Path]:
    """PDF builder for a pre-filtered receipt set (e.g. one calendar year)."""
    if not receipts:
        return []

    chunks = [
        receipts[i : i + RECEIPTS_PER_PDF]
        for i in range(0, len(receipts), RECEIPTS_PER_PDF)
    ]
    total_vat_all = sum(float(r.get("vat_amount") or 0) for r in receipts)
    total_parts = len(chunks)
    out_paths: list[Path] = []

    for part_idx, chunk in enumerate(chunks, 1):
        global_offset = (part_idx - 1) * RECEIPTS_PER_PDF
        suffix = f"_part{part_idx}_of_{total_parts}" if total_parts > 1 else ""
        out_path = output_dir / f"{name_prefix}{suffix}.pdf"
        await _write_pdf_chunk(
            chunk=chunk,
            employee_name=employee_name,
            part_idx=part_idx,
            total_parts=total_parts,
            global_offset=global_offset,
            grand_total_vat=total_vat_all,
            total_receipts=len(receipts),
            output_path=out_path,
        )
        out_paths.append(out_path)

    return out_paths


# ---- Per-receipt PDF page (soliq.uz data, no embedded photo) -----------
#
# The tax office now takes the originals and only needs a digital index,
# so each receipt is rendered as a single page using the structured
# soliq.uz data instead of the receipt photograph. Header block at the
# top, line-items table in the middle (if available), totals + soliq URL
# footer at the bottom. A KeepInFrame in 'shrink' mode auto-scales the
# whole block to fit one page even for receipts with many items.

_ACCENT = colors.HexColor("#0c4a8a")
_MUTED = colors.HexColor("#666666")
_RULE = colors.HexColor("#cfd6df")
_ZEBRA = colors.HexColor("#f3f5f8")

_STYLE_TITLE = ParagraphStyle(
    "title", fontName="Helvetica-Bold", fontSize=13, leading=15,
    textColor=colors.white,
)
_STYLE_SECTION = ParagraphStyle(
    "section", fontName="Helvetica-Bold", fontSize=10, leading=12,
    textColor=_ACCENT, spaceBefore=2, spaceAfter=2,
)
_STYLE_LABEL = ParagraphStyle(
    "label", fontName="Helvetica", fontSize=9, leading=11, textColor=_MUTED,
)
_STYLE_VALUE = ParagraphStyle(
    "value", fontName="Helvetica-Bold", fontSize=10, leading=12,
    textColor=colors.black,
)
_STYLE_BODY = ParagraphStyle(
    "body", fontName="Helvetica", fontSize=9, leading=11,
)
_STYLE_FOOTER = ParagraphStyle(
    "footer", fontName="Helvetica", fontSize=7, leading=9, textColor=_MUTED,
)
_STYLE_TOTAL = ParagraphStyle(
    "total", fontName="Helvetica-Bold", fontSize=13, leading=15,
    textColor=_ACCENT, alignment=2,  # right-aligned
)


def _esc(text) -> str:
    """HTML-escape a value for safe rendering inside a Paragraph, and
    normalize typographic punctuation that reportlab's default Helvetica
    can't render (Uzbek receipts use the curly apostrophe in words like
    "ko'chasi" and "o'zbek" — without normalization those glyphs show as
    a tofu rectangle)."""
    s = "" if text is None else str(text)
    s = (
        s.replace("‘", "'").replace("’", "'")
        .replace("“", '"').replace("”", '"')
        .replace("ʻ", "'").replace("ʼ", "'")
        .replace("′", "'").replace("″", '"')
        .replace("–", "-").replace("—", "-")
        .replace(" ", " ")
    )
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _build_receipt_page_story(r: dict, global_i: int, total_receipts: int,
                              frame_w: float) -> list:
    """Build the Platypus story for a single receipt's page. Returned
    flowables are wrapped in a KeepInFrame by the caller so the whole
    block shrinks to fit one page when items spill over."""
    vendor = _display_vendor(r) or "—"
    date = r.get("date") or "—"
    receipt_no = r.get("receipt_number") or "—"
    total = float(r.get("total_amount") or 0)
    vat = float(r.get("vat_amount") or 0)
    soliq_url = r.get("soliq_url") or r.get("raw_qr") or ""
    meta = r.get("soliq_meta") or {}
    items = r.get("soliq_items") or []

    story: list = []

    # Title bar: blue band with the page index and a soliq.uz badge.
    badge = "Verified by soliq.uz" if soliq_url else "Manual entry"
    title_table = Table(
        [[Paragraph(f"Receipt #{global_i} of {total_receipts}", _STYLE_TITLE),
          Paragraph(badge, _STYLE_TITLE)]],
        colWidths=[frame_w * 0.55, frame_w * 0.45],
    )
    title_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _ACCENT),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(title_table)
    story.append(Spacer(0, 6))

    # Header block — two columns of label/value pairs so the page reads
    # like the soliq.uz page itself (vendor first, then receipt-level
    # identifiers).
    header_rows = [
        ("Vendor", vendor),
        ("Date", date),
        ("Receipt #", receipt_no),
    ]
    if meta.get("time"):
        header_rows.append(("Time", meta["time"]))
    if meta.get("stir"):
        header_rows.append(("STIR / INN", meta["stir"]))
    if meta.get("terminal"):
        header_rows.append(("Terminal", meta["terminal"]))
    if meta.get("address"):
        header_rows.append(("Address", meta["address"]))
    if meta.get("cashier"):
        header_rows.append(("Cashier", meta["cashier"]))

    # Lay out as two columns of (label, value, label, value) so we use
    # the width efficiently and short fields don't waste a line.
    table_rows = []
    for i in range(0, len(header_rows), 2):
        left = header_rows[i]
        right = header_rows[i + 1] if i + 1 < len(header_rows) else ("", "")
        table_rows.append([
            Paragraph(_esc(left[0]), _STYLE_LABEL),
            Paragraph(_esc(left[1]), _STYLE_VALUE),
            Paragraph(_esc(right[0]), _STYLE_LABEL),
            Paragraph(_esc(right[1]), _STYLE_VALUE),
        ])
    header_table = Table(
        table_rows,
        colWidths=[
            frame_w * 0.13, frame_w * 0.37,
            frame_w * 0.13, frame_w * 0.37,
        ],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(header_table)
    story.append(Spacer(0, 6))

    # Items table — only rendered when the soliq scrape captured items.
    # Old receipts saved before this rewrite have an empty list and skip
    # straight to totals, which still looks fine.
    if items:
        story.append(Paragraph("Items", _STYLE_SECTION))
        item_rows = [[
            Paragraph("<b>#</b>", _STYLE_BODY),
            Paragraph("<b>Product</b>", _STYLE_BODY),
            Paragraph("<b>Qty</b>", _STYLE_BODY),
            Paragraph("<b>Price</b>", _STYLE_BODY),
            Paragraph("<b>VAT</b>", _STYLE_BODY),
            Paragraph("<b>Total</b>", _STYLE_BODY),
        ]]
        for idx, item in enumerate(items, 1):
            qty = float(item.get("qty") or 0)
            price = float(item.get("price") or 0)
            line_total = float(item.get("total") or 0)
            line_vat = float(item.get("vat") or 0)
            item_rows.append([
                Paragraph(str(idx), _STYLE_BODY),
                Paragraph(_esc(item.get("name") or ""), _STYLE_BODY),
                Paragraph(f"{qty:g}" if qty else "—", _STYLE_BODY),
                Paragraph(f"{price:,.2f}" if price else "—", _STYLE_BODY),
                Paragraph(f"{line_vat:,.2f}" if line_vat else "—", _STYLE_BODY),
                Paragraph(f"{line_total:,.2f}" if line_total else "—", _STYLE_BODY),
            ])
        items_table = Table(
            item_rows,
            colWidths=[
                frame_w * 0.05,
                frame_w * 0.49,
                frame_w * 0.08,
                frame_w * 0.13,
                frame_w * 0.11,
                frame_w * 0.14,
            ],
            repeatRows=1,
        )
        items_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _ACCENT),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _ZEBRA]),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, _ACCENT),
            ("LINEBELOW", (0, -1), (-1, -1), 0.5, _RULE),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(items_table)
        story.append(Spacer(0, 4))

    # Totals block: subtotal (only when total > vat so we don't print a
    # negative number on manual entries that lack the gross amount), VAT,
    # and the grand-total anchor.
    totals_rows = []
    if total > vat > 0:
        totals_rows.append([
            Paragraph("Subtotal", _STYLE_LABEL),
            Paragraph(f"{total - vat:,.2f} UZS", _STYLE_VALUE),
        ])
    totals_rows.append([
        Paragraph("VAT (QQS)", _STYLE_LABEL),
        Paragraph(f"{vat:,.2f} UZS", _STYLE_VALUE),
    ])
    if total > 0:
        totals_rows.append([
            Paragraph("<b>TOTAL</b>", _STYLE_SECTION),
            Paragraph(f"{total:,.2f} UZS", _STYLE_TOTAL),
        ])
    totals_table = Table(
        totals_rows,
        colWidths=[frame_w * 0.7, frame_w * 0.3],
    )
    totals_table.setStyle(TableStyle([
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LINEABOVE", (0, -1), (-1, -1), 0.5, _ACCENT),
        ("LINEBELOW", (0, -1), (-1, -1), 1.2, _ACCENT),
    ]))
    story.append(Spacer(0, 4))
    story.append(totals_table)

    if soliq_url:
        story.append(Spacer(0, 6))
        story.append(Paragraph(
            f"Source: <font color='#0c4a8a'>{_esc(soliq_url)}</font>",
            _STYLE_FOOTER,
        ))

    return story


async def _write_pdf_chunk(
    *,
    chunk: list[dict],
    employee_name: str,
    part_idx: int,
    total_parts: int,
    global_offset: int,
    grand_total_vat: float,
    total_receipts: int,
    output_path: Path,
) -> None:
    c = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4
    margin = 15 * mm

    # --- Cover / summary page ---
    title = "Tashkent Embassy VAT Refund — Receipt Package"
    if total_parts > 1:
        title += f"  (Part {part_idx} of {total_parts})"
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, height - margin - 10 * mm, title)
    c.setFont("Helvetica", 11)
    c.drawString(margin, height - margin - 18 * mm, f"Employee: {employee_name or '—'}")
    c.drawString(
        margin, height - margin - 24 * mm,
        f"Receipts in this file: {len(chunk)} (#{global_offset + 1}–#{global_offset + len(chunk)} of {total_receipts})",
    )
    chunk_total = sum(float(r.get("vat_amount") or 0) for r in chunk)
    c.drawString(margin, height - margin - 30 * mm, f"VAT in this file (UZS): {chunk_total:,.2f}")
    if total_parts > 1:
        c.drawString(margin, height - margin - 36 * mm, f"Grand total VAT across all parts: {grand_total_vat:,.2f} UZS")
        summary_y = height - margin - 51 * mm
    else:
        summary_y = height - margin - 45 * mm

    # Summary table for this chunk
    y = summary_y
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin, y, "#")
    c.drawString(margin + 10 * mm, y, "Date")
    c.drawString(margin + 35 * mm, y, "Vendor")
    c.drawString(margin + 100 * mm, y, "Receipt #")
    c.drawRightString(margin + 185 * mm, y, "VAT (UZS)")
    y -= 2 * mm
    c.setStrokeColorRGB(0.6, 0.6, 0.6)
    c.setLineWidth(0.3)
    c.line(margin, y, width - margin, y)
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    for i, r in enumerate(chunk, 1):
        if y < margin + 15 * mm:
            c.showPage()
            y = height - margin
            c.setFont("Helvetica", 9)
        c.drawString(margin, y, str(global_offset + i))
        c.drawString(margin + 10 * mm, y, str(r.get("date", "") or ""))
        vendor = _display_vendor(r)[:32]
        c.drawString(margin + 35 * mm, y, vendor)
        c.drawString(margin + 100 * mm, y, str(r.get("receipt_number", "") or "")[:18])
        vat = float(r.get("vat_amount") or 0)
        c.drawRightString(margin + 185 * mm, y, f"{vat:,.2f}")
        y -= 5 * mm
    y += 2 * mm
    c.line(margin, y, width - margin, y)
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin + 100 * mm, y, "SUBTOTAL" if total_parts > 1 else "TOTAL")
    c.drawRightString(margin + 185 * mm, y, f"{chunk_total:,.2f}")

    # --- One page per receipt, rendered from soliq.uz data ---
    frame_w = width - 2 * margin
    frame_h = height - 2 * margin - 8 * mm  # leave room for the page footer
    for i, r in enumerate(chunk, 1):
        global_i = global_offset + i
        c.showPage()
        try:
            story = _build_receipt_page_story(r, global_i, total_receipts, frame_w)
            keeper = KeepInFrame(frame_w, frame_h, story, mode="shrink")
            frame = Frame(
                margin, margin + 8 * mm, frame_w, frame_h,
                leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
                showBoundary=0,
            )
            frame.addFromList([keeper], c)
        except Exception:
            c.setFont("Helvetica-Oblique", 10)
            c.drawString(margin, height / 2, "[Receipt could not be rendered]")

        c.setFont("Helvetica", 8)
        c.setFillGray(0.5)
        c.drawString(margin, margin / 2, employee_name or "")
        c.drawRightString(width - margin, margin / 2, f"Receipt {global_i} of {total_receipts}")
        c.setFillGray(0)

    c.save()
